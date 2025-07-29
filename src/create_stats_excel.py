#!/usr/bin/env python3
"""
Poker Stats Excel Generator

Creates a clean, formatted Excel file with hero stats and player analysis.
Includes multiple sheets with different views of the data.
"""

import json
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
import os

class StatsExcelGenerator:
    def __init__(self, hero_name: str = "pmatheis"):
        self.hero_name = hero_name
        self.hero_stats = {}
        self.players_stats = {}
        
    def load_data(self, hero_stats_file: str, players_stats_file: str):
        """Load hero and player stats from JSON files"""
        try:
            with open(hero_stats_file, 'r') as f:
                self.hero_stats = json.load(f)
            print(f"Loaded hero stats from {hero_stats_file}")
            
            with open(players_stats_file, 'r') as f:
                self.players_stats = json.load(f)
            print(f"Loaded {len(self.players_stats)} player stats from {players_stats_file}")
            
        except FileNotFoundError as e:
            print(f"Error: {e}")
            sys.exit(1)
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON: {e}")
            sys.exit(1)
    
    def create_hero_overview_data(self):
        """Create hero overview data for summary sheet"""
        overview_data = [
            ["Metric", "Value"],
            ["Total Hands", f"{self.hero_stats['total_hands_played']:,}"],
            ["Net Result", f"${self.hero_stats['net_result']:.2f}"],
            ["Total Winnings", f"${self.hero_stats['total_winnings']:.2f}"],
            ["Total Invested", f"${self.hero_stats['total_invested']:.2f}"],
            ["ROI", f"{self.hero_stats['roi_percentage']:.1f}%"],
            ["BB/100 Hands", f"{self.hero_stats['bb_per_100_hands']:.2f}"],
            ["", ""],  # Spacer
            ["VPIP", f"{self.hero_stats['vpip_percentage']:.1f}%"],
            ["PFR", f"{self.hero_stats['pfr_percentage']:.1f}%"],
            ["3-Bet", f"{self.hero_stats['three_bet_percentage']:.1f}%"],
            ["Showdown Win Rate", f"{self.hero_stats['won_at_showdown_percentage']:.1f}%"],
            ["C-Bet", f"{self.hero_stats['cbet_percentage']:.1f}%"],
            ["Aggression Factor", f"{self.hero_stats['aggression_factor']:.2f}"],
            ["", ""],  # Spacer
            ["Total Sessions", f"{self.hero_stats['total_sessions']}"],
            ["Winning Sessions", f"{self.hero_stats['winning_sessions']}"],
            ["Session Win Rate", f"{self.hero_stats['session_win_rate_percentage']:.1f}%"],
            ["Best Session", f"${self.hero_stats['best_session']:.2f}"],
            ["Worst Session", f"${self.hero_stats['worst_session']:.2f}"],
            ["Average Session", f"${self.hero_stats['average_session']:.2f}"]
        ]
        return overview_data
    
    def create_players_dataframe(self):
        """Create DataFrame with all player stats"""
        players_data = []
        
        for player_name, stats in self.players_stats.items():
            player_row = {
                'Player': player_name,
                'Hands vs Hero': stats['hands_against_hero'],
                'Hero Net vs Player': stats['hero_net_vs_player'],
                'Player Net Result': stats['net_result'],
                'Player Total Winnings': stats['total_winnings'],
                'Player Total Invested': stats['total_invested'],
                'VPIP %': stats['vpip_percentage'],
                'PFR %': stats['pfr_percentage'],
                '3-Bet %': stats['three_bet_percentage'],
                'Showdown %': stats['went_to_showdown_percentage'],
                'Showdown Win %': stats['won_at_showdown_percentage'],
                'C-Bet %': stats['cbet_percentage'],
                'Fold vs Flop %': stats['fold_vs_flop_bet_percentage'],
                'AFq %': stats['afq_percentage'],
                'Turn CBet %': stats['turn_cbet_percentage'],
                'Fold vs Turn %': stats['fold_vs_turn_bet_percentage'],
                'River CBet %': stats['river_cbet_percentage'],
                'Fold vs River %': stats['fold_vs_river_bet_percentage'],
                'Donk Bet %': stats['donk_bet_percentage'],
                'Aggression Factor': stats['aggression_factor'],
                'Aggressive Actions': stats['total_aggressive_actions'],
                'Passive Actions': stats['total_passive_actions']
            }
            players_data.append(player_row)
        
        df = pd.DataFrame(players_data)
        return df.sort_values('Hands vs Hero', ascending=False)
    
    def apply_hero_overview_styling(self, ws, data):
        """Apply styling to hero overview sheet"""
        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Apply header styling
        for col in range(1, 3):  # A1:B1
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # Style data rows
        for row_idx, row_data in enumerate(data[1:], start=2):
            metric_cell = ws.cell(row=row_idx, column=1)
            value_cell = ws.cell(row=row_idx, column=2)
            
            # Bold metric names (except spacers)
            if row_data[0] != "":
                metric_cell.font = Font(bold=True)
            
            # Color positive/negative values
            if isinstance(row_data[1], str) and '$' in row_data[1] and '-' in row_data[1]:
                value_cell.font = Font(color="DC143C")  # Red for losses
            elif isinstance(row_data[1], str) and '$' in row_data[1] and '-' not in row_data[1] and row_data[1] != "$0.00":
                value_cell.font = Font(color="228B22")  # Green for profits
            
            # Align values to the right
            value_cell.alignment = Alignment(horizontal="right")
    
    def apply_dataframe_styling(self, ws, df):
        """Apply styling to dataframe sheets"""
        # Header styling
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styling
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        # Auto-fit columns and apply borders
        for idx, column in enumerate(df.columns, 1):
            column_letter = ws.cell(row=1, column=idx).column_letter
            
            # Set column width based on content
            max_length = max(
                len(str(column)),
                max([len(str(df[column].iloc[i])) for i in range(min(len(df), 10))])
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
            
            # Apply borders to all cells in column
            for row in range(1, len(df) + 2):
                ws.cell(row=row, column=idx).border = thin_border
        
        # Apply conditional formatting for profit/loss columns
        profit_columns = ['Hero Net vs Player', 'Player Net Result']
        for col_name in profit_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                
                # Green for positive values
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{len(df)+1}',
                    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"))
                )
                
                # Red for negative values
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{len(df)+1}',
                    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"))
                )
    
    def create_excel_file(self, output_filename: str = None):
        """Create the Excel file with multiple sheets"""
        if output_filename is None:
            output_filename = f"data/{self.hero_name}_poker_stats.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        ws_default = wb.active
        wb.remove(ws_default)
        
        # 1. Hero Overview Sheet
        ws_overview = wb.create_sheet("Hero Overview")
        overview_data = self.create_hero_overview_data()
        
        for row_data in overview_data:
            ws_overview.append(row_data)
        
        self.apply_hero_overview_styling(ws_overview, overview_data)
        
        # 2. Hero Details Sheet (all stats in table format)
        ws_hero_details = wb.create_sheet("Hero Details")
        hero_details_data = []
        
        for key, value in self.hero_stats.items():
            formatted_key = key.replace('_', ' ').title()
            if isinstance(value, float):
                if 'percentage' in key:
                    formatted_value = f"{value:.1f}%"
                elif any(word in key for word in ['result', 'winnings', 'invested', 'rake', 'session']):
                    formatted_value = f"${value:.2f}"
                else:
                    formatted_value = f"{value:.2f}"
            else:
                formatted_value = f"{value:,}" if isinstance(value, int) else str(value)
            
            hero_details_data.append([formatted_key, formatted_value])
        
        hero_df = pd.DataFrame(hero_details_data, columns=['Statistic', 'Value'])
        
        for r in dataframe_to_rows(hero_df, index=False, header=True):
            ws_hero_details.append(r)
        
        self.apply_dataframe_styling(ws_hero_details, hero_df)
        
        # 3. Top Players Analysis Sheet (filtered for significance)
        ws_top_players = wb.create_sheet("Top Players Analysis")
        players_df = self.create_players_dataframe()
        
        # Filter for players with significant sample size (50+ hands)
        significant_players = players_df[players_df['Hands vs Hero'] >= 50].head(20)
        
        for r in dataframe_to_rows(significant_players, index=False, header=True):
            ws_top_players.append(r)
        
        self.apply_dataframe_styling(ws_top_players, significant_players)
        
        # 4. All Players Sheet
        ws_all_players = wb.create_sheet("All Players")
        
        for r in dataframe_to_rows(players_df, index=False, header=True):
            ws_all_players.append(r)
        
        self.apply_dataframe_styling(ws_all_players, players_df)
        
        # 5. Profitability Analysis Sheet
        ws_profit_analysis = wb.create_sheet("Profitability Analysis")
        
        # Most profitable opponents
        most_profitable = players_df.nlargest(10, 'Hero Net vs Player')[['Player', 'Hands vs Hero', 'Hero Net vs Player']]
        most_profitable.columns = ['Most Profitable Opponents', 'Hands', 'Hero Profit']
        
        # Biggest losses
        biggest_losses = players_df.nsmallest(10, 'Hero Net vs Player')[['Player', 'Hands vs Hero', 'Hero Net vs Player']]
        biggest_losses.columns = ['Biggest Loss Opponents', 'Hands', 'Hero Loss']
        
        # Combine into one sheet with spacing
        profit_data = []
        profit_data.extend([list(most_profitable.columns)])
        for _, row in most_profitable.iterrows():
            profit_data.append([row['Most Profitable Opponents'], row['Hands'], f"${row['Hero Profit']:.2f}"])
        
        profit_data.append(['', '', ''])  # Spacer
        profit_data.append(['', '', ''])  # Spacer
        profit_data.extend([list(biggest_losses.columns)])
        for _, row in biggest_losses.iterrows():
            profit_data.append([row['Biggest Loss Opponents'], row['Hands'], f"${row['Hero Loss']:.2f}"])
        
        for row_data in profit_data:
            ws_profit_analysis.append(row_data)
        
        # Style the profitability sheet
        ws_profit_analysis.column_dimensions['A'].width = 20
        ws_profit_analysis.column_dimensions['B'].width = 10
        ws_profit_analysis.column_dimensions['C'].width = 12
        
        # Save the file
        wb.save(output_filename)
        print(f"Excel file saved to: {output_filename}")
        
        return output_filename

def main():
    if len(sys.argv) < 3:
        print("Usage: python3 create_stats_excel.py <hero_stats.json> <players_stats.json> [hero_name]")
        print("  Creates a comprehensive Excel file with poker statistics")
        sys.exit(1)
    
    hero_stats_file = sys.argv[1]
    players_stats_file = sys.argv[2]
    hero_name = sys.argv[3] if len(sys.argv) > 3 else "pmatheis"
    
    print(f"Creating Excel stats file for {hero_name}")
    
    generator = StatsExcelGenerator(hero_name)
    generator.load_data(hero_stats_file, players_stats_file)
    output_file = generator.create_excel_file()
    
    print(f"\nExcel file created successfully!")
    print(f"File contains the following sheets:")
    print("  - Hero Overview: Key performance metrics")
    print("  - Hero Details: Complete statistics table")
    print("  - Top Players Analysis: Players with 50+ hands")
    print("  - All Players: Complete opponent database")
    print("  - Profitability Analysis: Best and worst matchups")

if __name__ == "__main__":
    main() 